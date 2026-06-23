import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load test case metadata for lookup
def load_metadata_map():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_path = os.path.join(base_dir, "data", "test_cases.json")
    if os.path.exists(data_path):
        try:
            with open(data_path, "r", encoding="utf-8") as f:
                tcs = json.load(f)
                return {tc["id"]: tc for tc in tcs}
        except Exception:
            pass
    return {}

def enrich_results(test_results):
    tc_map = load_metadata_map()
    enriched = []
    for idx, r in enumerate(test_results, 1):
        name = r["name"]
        test_id = f"TC_GEN_{idx:03}"
        
        # Extract ID from name format: test_regression_case[TC_AUTH_001]
        if "[" in name and "]" in name:
            extracted_id = name.split("[")[-1].split("]")[0]
            if extracted_id in tc_map:
                test_id = extracted_id
                
        meta = tc_map.get(test_id, {})
        enriched.append({
            "id": test_id,
            "module": meta.get("module", "General"),
            "name": meta.get("name", name),
            "priority": meta.get("priority", "Medium"),
            "status": r["status"].upper(),
            "execution_time": r["duration"],
            "timestamp": r["timestamp"],
            "error": r.get("error", "")
        })
    return enriched

def apply_header_style(ws, title, cols_count):
    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    border_side = Side(style="thin", color="D5D8DC")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols_count)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    ws.sheet_view.showGridLines = True

def apply_table_headers(ws, headers):
    navy_fill = PatternFill(start_color="2A4D7C", end_color="2A4D7C", fill_type="solid")
    border_side = Side(style="thin", color="D5D8DC")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.row_dimensions[3].height = 26
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def write_test_rows(ws, tests, green_fill, red_fill, yellow_fill, thin_border):
    for r_idx, t in enumerate(tests, 4):
        ws.row_dimensions[r_idx].height = 20
        vals = [
            t["id"],
            t["module"],
            t["name"],
            t["priority"],
            t["status"],
            f"{t['execution_time']:.2f}",
            t["timestamp"],
            t.get("error", "")
        ]
        
        status = t["status"].lower()
        fill = green_fill if status == "passed" else red_fill if status == "failed" else yellow_fill
        
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            if c_idx == 5: # Status
                cell.fill = fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            elif c_idx in (1, 4, 6, 7):
                cell.alignment = Alignment(horizontal="center")

def autofit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

def generate_excel_report(test_results, output_path):
    """Generates the primary Automation_Test_Report.xlsx containing 7 worksheets."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    enriched = enrich_results(test_results)
    
    total = len(enriched)
    passed = [t for t in enriched if t["status"] == "PASSED"]
    failed = [t for t in enriched if t["status"] == "FAILED"]
    skipped = [t for t in enriched if t["status"] == "SKIPPED"]
    pass_pct = (len(passed) / total * 100) if total > 0 else 0
    
    wb = Workbook()
    
    green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    red_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    yellow_fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
    border_side = Side(style="thin", color="D5D8DC")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # ── Sheet 1: Executed Test Cases ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Executed Test Cases"
    apply_header_style(ws1, "Executed Test Cases - Mobile E2E", 8)
    headers = ["Test ID", "Module", "Test Name", "Priority", "Status", "Execution Time (s)", "Timestamp", "Trace/Errors"]
    apply_table_headers(ws1, headers)
    write_test_rows(ws1, enriched, green_fill, red_fill, yellow_fill, thin_border)
    autofit_columns(ws1)
    
    # ── Sheet 2: Passed Tests ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Passed Tests")
    apply_header_style(ws2, "Passed Test Cases Details", 7)
    apply_table_headers(ws2, headers[:7])
    write_test_rows(ws2, passed, green_fill, red_fill, yellow_fill, thin_border)
    autofit_columns(ws2)
    
    # ── Sheet 3: Failed Tests ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Failed Tests")
    apply_header_style(ws3, "Failed Test Cases Details", 8)
    apply_table_headers(ws3, headers)
    write_test_rows(ws3, failed, green_fill, red_fill, yellow_fill, thin_border)
    autofit_columns(ws3)
    
    # ── Sheet 4: Skipped Tests ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Skipped Tests")
    apply_header_style(ws4, "Skipped Test Cases Details", 8)
    apply_table_headers(ws4, headers)
    write_test_rows(ws4, skipped, green_fill, red_fill, yellow_fill, thin_border)
    autofit_columns(ws4)
    
    # ── Sheet 5: Execution Metrics ───────────────────────────────────────────
    ws5 = wb.create_sheet("Execution Metrics")
    apply_header_style(ws5, "High-Level Automation Execution Metrics", 2)
    ws5.cell(row=3, column=1, value="Metric Category").font = Font(bold=True)
    ws5.cell(row=3, column=2, value="Value").font = Font(bold=True)
    metrics_data = [
        ("Total Test Cases", total),
        ("Passed Tests Count", len(passed)),
        ("Failed Tests Count", len(failed)),
        ("Skipped Tests Count", len(skipped)),
        ("Overall Pass Percentage", f"{pass_pct:.2f}%"),
        ("Run Environment", "Android Emulator (Ubuntu KVM)"),
        ("Reporting Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ]
    for idx, (label, val) in enumerate(metrics_data, 4):
        ws5.cell(row=idx, column=1, value=label).border = thin_border
        cell_v = ws5.cell(row=idx, column=2, value=val)
        cell_v.border = thin_border
        cell_v.alignment = Alignment(horizontal="left")
        if label == "Overall Pass Percentage":
            cell_v.fill = green_fill if pass_pct >= 95 else red_fill
            cell_v.font = Font(bold=True)
    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 40
    
    # ── Sheet 6: Defect Summary ──────────────────────────────────────────────
    ws6 = wb.create_sheet("Defect Summary")
    apply_header_style(ws6, "Failed Tests & Defect Categorization Log", 5)
    ws6.row_dimensions[3].height = 24
    def_headers = ["Defect ID", "Failed Test ID", "Module", "Failure Reason", "Severity"]
    for col_idx, dh in enumerate(def_headers, 1):
        c = ws6.cell(row=3, column=col_idx, value=dh)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
        c.border = thin_border
        
    for r_idx, t in enumerate(failed, 4):
        err_msg = t.get("error", "No assertion details captured.")
        if "\n" in err_msg:
            err_msg = err_msg.split("\n")[0]
        ws6.cell(row=r_idx, column=1, value=f"DEF_{r_idx-3}").border = thin_border
        ws6.cell(row=r_idx, column=2, value=t["id"]).border = thin_border
        ws6.cell(row=r_idx, column=3, value=t["module"]).border = thin_border
        ws6.cell(row=r_idx, column=4, value=err_msg).border = thin_border
        ws6.cell(row=r_idx, column=5, value="High").border = thin_border
    autofit_columns(ws6)
    
    # ── Sheet 7: Pass Rate Summary ───────────────────────────────────────────
    ws7 = wb.create_sheet("Pass Rate Summary")
    apply_header_style(ws7, "Module-wise Pass Rate Distribution Summary", 5)
    apply_table_headers(ws7, ["Module Name", "Total Tests", "Passed", "Failed", "Pass Rate (%)"])
    
    # Group metrics by module
    module_data = {}
    for t in enriched:
        mod = t["module"]
        if mod not in module_data:
            module_data[mod] = {"total": 0, "passed": 0, "failed": 0}
        module_data[mod]["total"] += 1
        if t["status"] == "PASSED":
            module_data[mod]["passed"] += 1
        elif t["status"] == "FAILED":
            module_data[mod]["failed"] += 1
            
    for idx, (mod_name, data) in enumerate(module_data.items(), 4):
        m_pct = (data["passed"] / data["total"] * 100) if data["total"] > 0 else 0
        ws7.cell(row=idx, column=1, value=mod_name).border = thin_border
        ws7.cell(row=idx, column=2, value=data["total"]).border = thin_border
        ws7.cell(row=idx, column=3, value=data["passed"]).border = thin_border
        ws7.cell(row=idx, column=4, value=data["failed"]).border = thin_border
        
        c_rate = ws7.cell(row=idx, column=5, value=f"{m_pct:.2f}%")
        c_rate.border = thin_border
        c_rate.font = Font(bold=True)
        c_rate.fill = green_fill if m_pct >= 95 else red_fill
        
    autofit_columns(ws7)
    
    wb.save(output_path)
    print(f"[MobileExcelReporter] Primary report saved successfully to {output_path}")

def generate_passed_report(test_results, output_path):
    """Generates Passed_Test_Cases.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Passed Cases"
    apply_header_style(ws, "Passed E2E Test Cases Record", 7)
    headers = ["Test ID", "Module", "Test Name", "Priority", "Status", "Execution Time (s)", "Timestamp"]
    apply_table_headers(ws, headers)
    
    green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    border_side = Side(style="thin", color="D5D8DC")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    enriched = enrich_results(test_results)
    passed = [t for t in enriched if t["status"] == "PASSED"]
    
    write_test_rows(ws, passed, green_fill, None, None, thin_border)
    autofit_columns(ws)
    wb.save(output_path)
    print(f"[MobileExcelReporter] Passed report saved to {output_path}")

def generate_failed_report(test_results, output_path):
    """Generates Failed_Test_Cases.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Failed Cases"
    apply_header_style(ws, "Failed E2E Test Cases Record", 8)
    headers = ["Test ID", "Module", "Test Name", "Priority", "Status", "Execution Time (s)", "Timestamp", "Failure Reason"]
    apply_table_headers(ws, headers)
    
    red_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    border_side = Side(style="thin", color="D5D8DC")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    enriched = enrich_results(test_results)
    failed = [t for t in enriched if t["status"] == "FAILED"]
    
    write_test_rows(ws, failed, None, red_fill, None, thin_border)
    autofit_columns(ws)
    wb.save(output_path)
    print(f"[MobileExcelReporter] Failed report saved to {output_path}")

def generate_summary_report(test_results, output_path):
    """Generates Execution_Summary.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary Metrics"
    apply_header_style(ws, "Executive Execution Metrics Summary Dashboard", 5)
    apply_table_headers(ws, ["Build #", "Total Tests", "Passed", "Failed", "Overall Pass Rate"])
    
    enriched = enrich_results(test_results)
    total = len(enriched)
    passed_count = sum(1 for t in enriched if t["status"] == "PASSED")
    failed_count = sum(1 for t in enriched if t["status"] == "FAILED")
    pass_pct = (passed_count / total * 100) if total > 0 else 0
    
    green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    red_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    border_side = Side(style="thin", color="D5D8DC")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.row_dimensions[4].height = 24
    ws.cell(row=4, column=1, value=f"#{os.getenv('GITHUB_RUN_NUMBER', '1')}").border = thin_border
    ws.cell(row=4, column=2, value=total).border = thin_border
    ws.cell(row=4, column=3, value=passed_count).border = thin_border
    ws.cell(row=4, column=4, value=failed_count).border = thin_border
    
    c_rate = ws.cell(row=4, column=5, value=f"{pass_pct:.2f}%")
    c_rate.border = thin_border
    c_rate.font = Font(bold=True)
    c_rate.fill = green_fill if pass_pct >= 95 else red_fill
    c_rate.alignment = Alignment(horizontal="center")
    
    autofit_columns(ws)
    wb.save(output_path)
    print(f"[MobileExcelReporter] Summary report saved to {output_path}")
